"""Camada de acesso a dados: única responsável por abrir o arquivo
PLANILHA CONTROLE DE GASTOS.xlsx, ler células cruas e gravar alterações em
disco. Não contém nenhuma regra de negócio — apenas I/O.
"""
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from api import config


class ExpenseRepository:
    def __init__(self, path: Optional[Path] = None):
        self._path = path or config.PLANILHA_PATH
        self._workbook = None

    def _ensure_loaded(self) -> None:
        if self._workbook is None:
            if not self._path.exists():
                raise FileNotFoundError(
                    f"Planilha não encontrada em: {self._path}"
                )
            self._workbook = openpyxl.load_workbook(self._path)

    @property
    def workbook(self):
        self._ensure_loaded()
        return self._workbook

    def get_sheet(self, sheet_name: str) -> Worksheet:
        wb = self.workbook
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Aba '{sheet_name}' não existe na planilha.")
        return wb[sheet_name]

    def get_or_create_sheet(self, sheet_name: str) -> Worksheet:
        wb = self.workbook
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        return wb[sheet_name]

    def get_cell(self, sheet_name: str, row: int, col: int):
        return self.get_sheet(sheet_name).cell(row=row, column=col).value

    def set_cell(self, sheet_name: str, row: int, col: int, value) -> None:
        # openpyxl trata value=None como "nenhum valor informado" ao usar o
        # kwarg de .cell(...), o que impediria limpar uma célula. Atribuir
        # .value diretamente garante que None realmente limpe a célula.
        cell = self.get_sheet(sheet_name).cell(row=row, column=col)
        cell.value = value

    def iter_rows(
        self,
        sheet_name: str,
        min_row: int = 1,
        max_row: Optional[int] = None,
        min_col: int = 1,
        max_col: Optional[int] = None,
    ) -> Iterable[tuple[Cell, ...]]:
        sheet = self.get_sheet(sheet_name)
        yield from sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        )

    def dimensions(self, sheet_name: str) -> tuple[int, int]:
        sheet = self.get_sheet(sheet_name)
        return sheet.max_row, sheet.max_column

    def save(self) -> None:
        self.workbook.save(self._path)

    def reload(self) -> None:
        self._workbook = None
